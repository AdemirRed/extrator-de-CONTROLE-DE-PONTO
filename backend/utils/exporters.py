"""
Exportadores: Excel, CSV, JSON, PDF.
"""

import io
import json
import csv
from datetime import datetime
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

from .calculations import minutes_to_str

EXPORTS_DIR = Path(__file__).parent.parent.parent / "exports"
EXPORTS_DIR.mkdir(exist_ok=True)


# ─── helpers ────────────────────────────────────────────────────────────────

def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color: str):
    return PatternFill("solid", fgColor=hex_color)


# ─── Excel ──────────────────────────────────────────────────────────────────

def export_excel(funcionario: dict, registros: list[dict], mes: int, ano: int) -> Path:
    MESES = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
             "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{MESES[mes-1]} {ano}"

    # cabeçalho
    ws.merge_cells("A1:I1")
    ws["A1"] = f"CARTÃO PONTO — {funcionario['nome'].upper()} — {MESES[mes-1].upper()} / {ano}"
    ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A1"].fill = _fill("0D1117")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # headers tabela
    headers = ["DIA","ENTRADA","IN. INTERVALO","FM. INTERVALO","SAÍDA","TOTAL","NOTURNO","PREVISTO","SALDO"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = _fill("161B22")
        cell.alignment = Alignment(horizontal="center")
        cell.border = _border()

    col_widths = [6, 10, 13, 13, 10, 10, 10, 10, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # linhas
    for i, reg in enumerate(registros, 3):
        vals = [
            reg.get("data","")[-2:],
            reg.get("entrada","") or "",
            reg.get("inicio_intervalo","") or "",
            reg.get("fim_intervalo","") or "",
            reg.get("saida","") or "",
            minutes_to_str(reg.get("total_min")),
            minutes_to_str(reg.get("noturno_min")),
            minutes_to_str(reg.get("previsto_min")),
            minutes_to_str(reg.get("saldo_min")),
        ]
        saldo = reg.get("saldo_min") or 0
        row_fill = None
        if saldo > 0:
            row_fill = _fill("0D2818")
        elif saldo < 0:
            row_fill = _fill("2D1515")

        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.alignment = Alignment(horizontal="center")
            cell.border = _border("hair")
            cell.font = Font(size=9)
            if row_fill:
                cell.fill = row_fill

    # totais
    tot_row = len(registros) + 3
    total_h  = sum(r.get("total_min") or 0 for r in registros)
    total_n  = sum(r.get("noturno_min") or 0 for r in registros)
    total_p  = sum(r.get("previsto_min") or 0 for r in registros)
    total_s  = sum(r.get("saldo_min") or 0 for r in registros)

    ws.cell(row=tot_row, column=1, value="TOTAL").font = Font(bold=True)
    for c, v in zip(range(6, 10), [
        minutes_to_str(total_h), minutes_to_str(total_n),
        minutes_to_str(total_p), minutes_to_str(total_s)
    ]):
        cell = ws.cell(row=tot_row, column=c, value=v)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _fill("161B22")
        cell.alignment = Alignment(horizontal="center")

    fname = EXPORTS_DIR / f"{funcionario['nome'].replace(' ','_')}_{ano}_{mes:02d}.xlsx"
    wb.save(fname)
    return fname


# ─── CSV ────────────────────────────────────────────────────────────────────

def export_csv(funcionario: dict, registros: list[dict], mes: int, ano: int) -> Path:
    fname = EXPORTS_DIR / f"{funcionario['nome'].replace(' ','_')}_{ano}_{mes:02d}.csv"
    with open(fname, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "data","entrada","inicio_intervalo","fim_intervalo","saida",
            "total","noturno","previsto","saldo","feriado"
        ])
        writer.writeheader()
        for r in registros:
            writer.writerow({
                "data":              r.get("data",""),
                "entrada":           r.get("entrada","") or "",
                "inicio_intervalo":  r.get("inicio_intervalo","") or "",
                "fim_intervalo":     r.get("fim_intervalo","") or "",
                "saida":             r.get("saida","") or "",
                "total":             minutes_to_str(r.get("total_min")),
                "noturno":           minutes_to_str(r.get("noturno_min")),
                "previsto":          minutes_to_str(r.get("previsto_min")),
                "saldo":             minutes_to_str(r.get("saldo_min")),
                "feriado":           r.get("feriado", 0),
            })
    return fname


# ─── JSON ───────────────────────────────────────────────────────────────────

def export_json(funcionario: dict, registros: list[dict], mes: int, ano: int) -> Path:
    fname = EXPORTS_DIR / f"{funcionario['nome'].replace(' ','_')}_{ano}_{mes:02d}.json"
    payload = {
        "funcionario": funcionario,
        "periodo": {"mes": mes, "ano": ano},
        "gerado_em": datetime.now().isoformat(),
        "registros": registros,
    }
    fname.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return fname
