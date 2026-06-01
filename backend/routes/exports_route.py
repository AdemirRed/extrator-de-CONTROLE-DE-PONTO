"""
Exportação de registros: Excel, CSV, JSON.
"""

from fastapi import APIRouter, Query, HTTPException
from fastapi.responses import FileResponse

from backend.database import get_connection
from backend.utils.exporters import export_excel, export_csv, export_json
from backend.utils.calculations import minutes_to_str

router = APIRouter(prefix="/api/exports", tags=["exports"])


def _get_data(fid: int, ano: int, mes: int):
    conn = get_connection()
    func = conn.execute("SELECT * FROM funcionarios WHERE id=?", (fid,)).fetchone()
    if not func:
        conn.close()
        raise HTTPException(404, "Funcionário não encontrado")
    regs = conn.execute(
        """SELECT * FROM registros
           WHERE funcionario_id=? AND data LIKE ?
           ORDER BY data""",
        (fid, f"{ano}-{mes:02d}-%"),
    ).fetchall()
    conn.close()
    return dict(func), [dict(r) for r in regs]


@router.get("/excel")
def excel(
    funcionario_id: int = Query(...),
    ano: int = Query(...),
    mes: int = Query(...),
):
    func, regs = _get_data(funcionario_id, ano, mes)
    path = export_excel(func, regs, mes, ano)
    return FileResponse(
        path=str(path),
        filename=path.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.get("/csv")
def csv_export(
    funcionario_id: int = Query(...),
    ano: int = Query(...),
    mes: int = Query(...),
):
    func, regs = _get_data(funcionario_id, ano, mes)
    path = export_csv(func, regs, mes, ano)
    return FileResponse(path=str(path), filename=path.name, media_type="text/csv")


@router.get("/json")
def json_export(
    funcionario_id: int = Query(...),
    ano: int = Query(...),
    mes: int = Query(...),
):
    func, regs = _get_data(funcionario_id, ano, mes)
    path = export_json(func, regs, mes, ano)
    return FileResponse(path=str(path), filename=path.name, media_type="application/json")


@router.get("/todos/excel")
def excel_todos(ano: int = Query(...), mes: int = Query(...)):
    """Exporta todos os funcionários em abas separadas do mesmo Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from pathlib import Path

    EXPORTS_DIR = Path(__file__).parent.parent.parent / "exports"
    MESES = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
             "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]

    conn = get_connection()
    funcs = conn.execute(
        "SELECT * FROM funcionarios WHERE ativo=1 ORDER BY nome"
    ).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for func in funcs:
        func = dict(func)
        conn2 = get_connection()
        regs = conn2.execute(
            "SELECT * FROM registros WHERE funcionario_id=? AND data LIKE ? ORDER BY data",
            (func["id"], f"{ano}-{mes:02d}-%"),
        ).fetchall()
        conn2.close()
        regs = [dict(r) for r in regs]

        ws = wb.create_sheet(func["nome"][:31])
        ws["A1"] = f"{func['nome'].upper()} — {MESES[mes-1]} {ano}"
        ws["A1"].font = Font(bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor="0D1117")
        ws.merge_cells("A1:I1")

        headers = ["DIA","ENTRADA","IN. INT","FM. INT","SAÍDA","TOTAL","NOTURNO","PREVISTO","SALDO"]
        s = Side(style="thin")
        bdr = Border(left=s,right=s,top=s,bottom=s)
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=c, value=h)
            cell.font = Font(bold=True, color="FFFFFF", size=9)
            cell.fill = PatternFill("solid", fgColor="161B22")
            cell.alignment = Alignment(horizontal="center")
            cell.border = bdr

        for i, reg in enumerate(regs, 3):
            vals = [
                reg.get("data","")[-2:],
                reg.get("entrada","") or "",
                reg.get("inicio_intervalo","") or "",
                reg.get("fim_intervalo","") or "",
                reg.get("saida","") or "",
                minutes_to_str(reg.get("total_min")),
                minutes_to_str(reg.get("noturno_min")),
                minutes_to_str(reg.get("previsto_min")),
                minutes_to_str(reg.get("saldo_min")),
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=i, column=c, value=v)
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(size=9)

    fname = EXPORTS_DIR / f"todos_{ano}_{mes:02d}.xlsx"
    wb.save(fname)
    return FileResponse(
        path=str(fname), filename=fname.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
